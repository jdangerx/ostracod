# coding: utf-8
from functools import reduce

import gspread
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl


def clean(string):
    if string == '' or string is None:
        return None
    return str(string).lower().strip()


def make_dict_from_cells(cells):
    headers, *rows = cells
    return make_dicts_from_rows(headers, rows)


def make_dicts_from_rows(headers, rows):
    headers_clean = [clean(h) for h in headers]
    rows_clean = [[clean(c) for c in r] for r in rows]
    return {clean(row[0]): dict(zip(headers_clean[1:], row[1:]))
            for row in rows_clean}


def merge_dicts(dict_of_dicts):
    all_keys = reduce(lambda m, n: m & n,
                      (set(d.keys()) for d in dict_of_dicts.values()))
    return {key: {clean(dict_name): dict_val.get(key, None)
                  for dict_name, dict_val in dict_of_dicts.items()} for key in all_keys}


def munge_species(traits, comments):
    species = {s: merge_dicts({"value": traits[s],
                               "comments": comments[s]})
               for s in traits}
    return species


def from_xlsx(filepath):
    wb = openpyxl.load_workbook(filepath)

    traits_sheet = wb.get_sheet_by_name("Traits")
    comments_sheet = wb.get_sheet_by_name("Comments")
    codings_sheet = wb.get_sheet_by_name("Trait codings")

    trait_names, *species_info = [[c.value for c in r]
                                  for r in traits_sheet["A3":"U{:d}".format(traits_sheet.max_row)]]
    comment_trait_names, *species_comments = [[c.value for c in r]
                                              for r in comments_sheet["A3":"U{:d}".format(comments_sheet.max_row)]]
    traits = make_dicts_from_rows(trait_names, species_info)
    comments = make_dicts_from_rows(comment_trait_names, species_comments)

    species = {s: merge_dicts({"value": traits[s], "comments": comments[s]}) for s in traits}

    coding_names, *coding_info = [[c.value for c in r]
                                  for r in codings_sheet["A1":"E{:d}".format(codings_sheet.max_row)]]
    codings = make_dicts_from_rows(coding_names, coding_info)
    return species, codings


def from_gsheet(keyfile_path, sheet_name):
    scope = ['https://spreadsheets.google.com/feeds']
    creds = ServiceAccountCredentials.from_json_keyfile_name(keyfile_path, scope)
    gc = gspread.authorize(creds)
    spreadsheet = gc.open(sheet_name)
    traits = spreadsheet.worksheet('Traits').get_all_values()
    comments = spreadsheet.worksheet('Comments').get_all_values()
    trait_codings = spreadsheet.worksheet('Trait codings').get_all_values()
    species = munge_species(
        make_dict_from_cells(traits),
        make_dict_from_cells(comments))
    return species, make_dict_from_cells(trait_codings)
